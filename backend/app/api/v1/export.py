"""Company data export API.

Builds a multi-sheet Excel (.xlsx) workbook for a single company — its stored
analysis plus every derived record (competitors, ICPs, personas, outreach
assets, market gaps, social intelligence) — and returns it as a file download.
Powers the dashboard "Export" button.

Endpoint:
  GET /export/company/{company_id}/xlsx — download the company's workbook.
"""

from __future__ import annotations

import json
import re
import uuid
from datetime import datetime
from io import BytesIO
from typing import Any

from fastapi import APIRouter, HTTPException, status
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.core.dependencies import DbSession
from app.core.logging import get_logger
from app.db.repositories import (
    CompanyRepository,
    CompetitorRepository,
    ICPRepository,
    MarketGapRepository,
    OutreachRepository,
    PersonaRepository,
    SocialContactRepository,
    SocialProfileRepository,
)

router = APIRouter(prefix="/export", tags=["Export"])
logger = get_logger(__name__)

# MIME type for an .xlsx (OOXML spreadsheet) document.
_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------


def _cell(value: Any) -> str:
    """Render any JSON-ish value as a single spreadsheet cell string.

    Nested dicts/lists (common in the JSONB columns) are serialised to compact
    JSON so they survive the export without needing one column per nested key.
    """
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _autosize(ws: Worksheet, max_width: int = 80) -> None:
    """Roughly fit each column to its widest cell, clamped to a sane range."""
    for column_cells in ws.columns:
        longest = max((len(str(c.value or "")) for c in column_cells), default=0)
        letter = column_cells[0].column_letter
        ws.column_dimensions[letter].width = min(max_width, max(12, longest + 2))


def _records_sheet(wb: Workbook, title: str, rows: list[dict]) -> None:
    """Add a sheet whose columns are the ordered union of keys across ``rows``."""
    ws = wb.create_sheet(title=title)
    if not rows:
        ws["A1"] = "No data"
        return
    keys: list[str] = []
    for row in rows:
        for key in row:
            if key not in keys:
                keys.append(key)
    ws.append(keys)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([_cell(row.get(k)) for k in keys])
    _autosize(ws)


# ---------------------------------------------------------------------------
# Endpoint
# ---------------------------------------------------------------------------


@router.get(
    "/company/{company_id}/xlsx",
    summary="Export all of a company's data as an Excel workbook",
)
async def export_company_xlsx(company_id: uuid.UUID, db: DbSession) -> Response:
    """GET /export/company/{company_id}/xlsx

    Gathers every record tied to the company and returns a single .xlsx file
    with one sheet per data type.
    """
    company = await CompanyRepository(db).get_by_id(company_id)
    if company is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Company not found"
        )

    competitors = await CompetitorRepository(db).get_by_company(company_id)
    icps = await ICPRepository(db).get_by_company(company_id)
    personas = await PersonaRepository(db).get_by_company(company_id)
    outreach = await OutreachRepository(db).get_by_company(company_id)
    gaps = await MarketGapRepository(db).get_by_company(company_id)
    profiles = await SocialProfileRepository(db).get_by_company(company_id)
    contacts = await SocialContactRepository(db).get_by_company(company_id)

    wb = Workbook()

    # --- Company overview (field/value layout) ---
    overview = wb.active
    overview.title = "Company"
    overview.append(["Field", "Value"])
    for cell in overview[1]:
        cell.font = Font(bold=True)
    overview.append(["Name", company.name])
    overview.append(["Industry", company.industry])
    overview.append(["Created", _cell(company.created_at)])
    for key, value in (company.input_data or {}).items():
        overview.append([f"input.{key}", _cell(value)])
    for key, value in (company.analysis or {}).items():
        overview.append([f"analysis.{key}", _cell(value)])
    _autosize(overview)

    # --- Derived-record sheets ---
    _records_sheet(
        wb,
        "Competitors",
        [
            {
                "Name": c.name,
                "Website": c.website,
                "Category": c.category,
                "Positioning": c.positioning,
                "Relevance": c.relevance_score,
                "Web Data": c.web_data,
                "Clean Data": c.clean_data,
                "Created": c.created_at,
            }
            for c in competitors
        ],
    )
    _records_sheet(
        wb,
        "ICPs",
        [
            {"ICP #": i + 1, "Created": icp.created_at, **(icp.profile_data or {})}
            for i, icp in enumerate(icps)
        ],
    )
    _records_sheet(
        wb,
        "Personas",
        [
            {
                "Persona #": i + 1,
                "ICP ID": str(p.icp_id),
                "Created": p.created_at,
                **(p.persona_data or {}),
            }
            for i, p in enumerate(personas)
        ],
    )
    _records_sheet(
        wb,
        "Outreach",
        [
            {
                "Channel": o.channel,
                "Persona ID": str(o.persona_id),
                "Open Rate": o.open_rate,
                "Reply Rate": o.reply_rate,
                "Conversion Rate": o.conversion_rate,
                "Created": o.created_at,
                **(o.content or {}),
            }
            for o in outreach
        ],
    )
    _records_sheet(
        wb,
        "Market Gaps",
        [
            {
                "Gap Type": g.gap_type,
                "Confidence": g.confidence_score,
                "Created": g.created_at,
                **(g.gap_data or {}),
            }
            for g in gaps
        ],
    )
    _records_sheet(
        wb,
        "Social Profiles",
        [
            {
                "Subject": p.subject_name,
                "Subject Type": p.subject_type,
                "Platform": p.platform,
                "Profile Type": p.profile_type,
                "URL": p.url,
                "Success": p.success,
                "Data": p.data,
                "Created": p.created_at,
            }
            for p in profiles
        ],
    )
    _records_sheet(
        wb,
        "Social Contacts",
        [
            {
                "Kind": c.kind,
                "Value": c.value,
                "Title": c.title,
                "Subject": c.subject_name,
                "URL": c.url,
                "Source Page": c.source_page,
                "Created": c.created_at,
            }
            for c in contacts
        ],
    )

    buffer = BytesIO()
    wb.save(buffer)

    # Filename slug — strip anything unsafe for a Content-Disposition header.
    slug = re.sub(r"[^A-Za-z0-9_-]+", "_", company.name).strip("_") or "company"
    filename = f"sellna_{slug}_export.xlsx"

    logger.info("api.export.company_xlsx", company_id=str(company_id))
    return Response(
        content=buffer.getvalue(),
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
